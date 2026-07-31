#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import sys
import textwrap
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas


DEFAULT_INPUTS: dict[str, Any] = {
    "projectName": "Bridge #42 — Rural Connect",
    "streamName": "Kaveri Tributary",
    "location": "District XYZ",
    "date": "2026-07-26",
    "catchmentArea": 12.5,
    "runoffCoefficient": 0.45,
    "rainfallIntensity": 80.0,
    "surplusWeirLength": 15.0,
    "heightOfFallWeir": 0.6,
    "streamAreaHFL": 45.0,
    "meanVelocityHFL": 1.2,
    "customDesignDischarge": None,
    "hfl": 102.5,
    "gl": 100.0,
    "rtl": 101.2,
    "numVents": 4,
    "ventWidth": 1.5,
    "ventHeight": 0.9,
    "approachVelocity": 1.2,
    "siltFactor": 1.0,
    "cdVent": 0.9,
    "deckWidth": 4.5,
    "deckSpan": 2.0,
    "deckThickness": 0.25,
    "numSpans": 6,
    "liveLoadType": "IRC Class A",
    "waterDensity": 1000.0,
    "concreteDensity": 2500.0,
    "dragCoefficient": 2.0,
    "siltLoadDeck": 1.2,
}

FIELD_TYPES: dict[str, str] = {
    "projectName": "string",
    "streamName": "string",
    "location": "string",
    "date": "string",
    "catchmentArea": "number",
    "runoffCoefficient": "number",
    "rainfallIntensity": "number",
    "surplusWeirLength": "number",
    "heightOfFallWeir": "number",
    "streamAreaHFL": "number",
    "meanVelocityHFL": "number",
    "customDesignDischarge": "nullable-number",
    "hfl": "number",
    "gl": "number",
    "rtl": "number",
    "numVents": "number",
    "ventWidth": "number",
    "ventHeight": "number",
    "approachVelocity": "number",
    "siltFactor": "number",
    "cdVent": "number",
    "deckWidth": "number",
    "deckSpan": "number",
    "deckThickness": "number",
    "numSpans": "number",
    "liveLoadType": "string",
    "waterDensity": "number",
    "concreteDensity": "number",
    "dragCoefficient": "number",
    "siltLoadDeck": "number",
}


def coerce_value(key: str, value: Any) -> Any:
    field_type = FIELD_TYPES.get(key, "string")
    if field_type == "string":
        return "" if value is None else str(value)
    raw = "" if value is None else str(value).strip()
    if field_type == "nullable-number":
        if raw.lower() in {"", "auto", "null", "none", "default"}:
            return None
        return float(raw)
    if field_type == "number":
        return float(raw)
    return value


def parse_spreadsheet_rows(rows: list[list[Any]]) -> dict[str, Any]:
    patch: dict[str, Any] = {}
    for row in rows:
        if not row:
            continue
        key = str(row[0]).strip()
        if not key or key.lower() == "variable":
            continue
        if key not in FIELD_TYPES:
            continue
        raw_value = row[1] if len(row) > 1 else ""
        patch[key] = coerce_value(key, raw_value)
    return patch


def load_inputs(input_path: Path | None) -> dict[str, Any]:
    data = dict(DEFAULT_INPUTS)
    if input_path is None:
        return data

    suffix = input_path.suffix.lower()
    patch: dict[str, Any] = {}
    if suffix == ".json":
        payload = json.loads(input_path.read_text(encoding="utf-8"))
        for key, value in payload.items():
            if key in FIELD_TYPES:
                patch[key] = coerce_value(key, value)
    elif suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        patch = parse_spreadsheet_rows(rows)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(input_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        patch = parse_spreadsheet_rows(rows)
    else:
        raise ValueError(f"Unsupported input file type: {input_path.suffix}")

    data.update(patch)
    data["numVents"] = int(round(float(data["numVents"])))
    data["numSpans"] = int(round(float(data["numSpans"])))
    return data


def compute_results(inputs: dict[str, Any]) -> dict[str, float | str | bool]:
    q_rational = (inputs["runoffCoefficient"] * inputs["rainfallIntensity"] * inputs["catchmentArea"]) / 3.6
    q_weir = 1.705 * inputs["surplusWeirLength"] * math.pow(inputs["heightOfFallWeir"], 1.5)
    q_velocity = inputs["streamAreaHFL"] * inputs["meanVelocityHFL"]

    q_design = q_rational
    governing_method = "Rational Method"
    if q_weir > q_design:
        q_design = q_weir
        governing_method = "Weir Formula"
    if q_velocity > q_design:
        q_design = q_velocity
        governing_method = "Area-Velocity"

    design_discharge = inputs["customDesignDischarge"] if inputs["customDesignDischarge"] is not None else q_design

    a_vent = inputs["numVents"] * inputs["ventWidth"] * inputs["ventHeight"]
    app_vel = inputs["approachVelocity"] if inputs["approachVelocity"] > 0 else 0.01
    effective_width = design_discharge / app_vel
    depth_rtl = max(0.0, inputs["rtl"] - inputs["gl"])
    a_rtl = depth_rtl * effective_width
    depth_hfl = max(0.0, inputs["hfl"] - inputs["gl"])
    a_hfl = depth_hfl * effective_width
    pct_obs_rtl = (1 - (a_vent / a_rtl)) * 100 if a_rtl > 0 else 0.0
    pct_obs_hfl = (1 - (a_vent / a_hfl)) * 100 if a_hfl > 0 else 0.0
    pass_rtl = pct_obs_rtl < 70
    pass_hfl = pct_obs_hfl < 30
    velocity_hfl = design_discharge / a_hfl if a_hfl > 0 else 0.0
    h_afflux = 0.0
    if a_vent > 0 and a_hfl > a_vent:
        h_afflux = (math.pow(velocity_hfl, 2) / 17.88 + 0.015) * (math.pow(a_hfl / a_vent, 2) - 1)
    lacey_perimeter = 4.75 * math.sqrt(design_discharge)
    sf = inputs["siltFactor"] if inputs["siltFactor"] > 0 else 0.1
    lacey_scour_depth = 0.473 * math.pow(design_discharge / sf, 1 / 3)
    max_scour_depth = 1.27 * lacey_scour_depth
    fbl = inputs["hfl"] - max_scour_depth
    recommended_depth = inputs["gl"] - fbl
    scour_safe = fbl < (inputs["gl"] - 0.5)
    w_self = (inputs["concreteDensity"] * 9.81 * inputs["deckWidth"] * inputs["deckSpan"] * inputs["deckThickness"]) / 1000
    w_silt = inputs["siltLoadDeck"] * inputs["deckWidth"] * inputs["deckSpan"]
    w_live = 700 if inputs["liveLoadType"] == "IRC Class AA" else 554
    total_vertical_load = w_self + w_silt + (w_live / inputs["numSpans"])
    f_drag = (inputs["dragCoefficient"] * 0.5 * inputs["waterDensity"] * math.pow(velocity_hfl, 2) * (inputs["deckWidth"] * inputs["deckThickness"])) / 1000
    f_uplift = (inputs["waterDensity"] * 9.81 * (inputs["deckWidth"] * inputs["deckSpan"] * inputs["deckThickness"])) / 1000
    f_anchor = f_uplift - w_self
    f_drag_total = f_drag * inputs["numSpans"]
    return {
        "qRational": q_rational,
        "qWeir": q_weir,
        "qVelocity": q_velocity,
        "qDesign": q_design,
        "governingMethod": governing_method,
        "designDischarge": design_discharge,
        "aVent": a_vent,
        "effectiveWidth": effective_width,
        "aRTL": a_rtl,
        "aHFL": a_hfl,
        "pctObsRTL": pct_obs_rtl,
        "pctObsHFL": pct_obs_hfl,
        "passRTL": pass_rtl,
        "passHFL": pass_hfl,
        "velocityHFL": velocity_hfl,
        "hAfflux": h_afflux,
        "laceyPerimeter": lacey_perimeter,
        "laceyScourDepth": lacey_scour_depth,
        "maxScourDepth": max_scour_depth,
        "fbl": fbl,
        "recommendedDepth": recommended_depth,
        "scourSafe": scour_safe,
        "wSelf": w_self,
        "wSilt": w_silt,
        "wLive": w_live,
        "totalVerticalLoad": total_vertical_load,
        "fDrag": f_drag,
        "fUplift": f_uplift,
        "fAnchor": f_anchor,
        "fDragTotal": f_drag_total,
    }


def f(num: float, digits: int = 3) -> str:
    return f"{num:.{digits}f}"


def replacement_map(inputs: dict[str, Any], results: dict[str, Any]) -> dict[str, str]:
    return {
        "B.T to the R/f KB Road to P.Bheemavaram": inputs["projectName"],
        "6.235": f(inputs["hfl"]),
        "3.965": f(inputs["gl"]),
        "5.645": f(inputs["rtl"]),
        "45.75": f(results["designDischarge"], 2),
        "26.21": f(results["aVent"], 2),
        "29.07": f(results["pctObsHFL"], 2),
        "39.40": f(results["pctObsRTL"], 2),
        "0.131": f(results["hAfflux"], 3),
        "2.63": f(results["maxScourDepth"], 2),
        "2.315": f(results["fbl"], 3),
    }


def apply_replacements(text: str, replacements: dict[str, str]) -> str:
    updated = text
    for old, new in replacements.items():
        updated = updated.replace(old, new)
    return updated


def draw_header(pdf: canvas.Canvas, title: str, subtitle: str, page_no: int, total_pages: int) -> None:
    width, height = A4
    pdf.setStrokeColorRGB(0.15, 0.19, 0.27)
    pdf.setLineWidth(0.8)
    pdf.rect(12 * mm, 12 * mm, width - 24 * mm, height - 24 * mm)
    pdf.setFillColorRGB(0.12, 0.18, 0.30)
    pdf.rect(12 * mm, height - 28 * mm, width - 24 * mm, 16 * mm, fill=1, stroke=0)
    pdf.setFillColorRGB(1, 1, 1)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(18 * mm, height - 20 * mm, title[:90])
    pdf.setFont("Helvetica", 8.5)
    pdf.drawRightString(width - 18 * mm, height - 20 * mm, f"Page {page_no} / {total_pages}")
    pdf.setFillColorRGB(0.24, 0.28, 0.35)
    pdf.setFont("Helvetica", 8)
    pdf.drawString(18 * mm, height - 32 * mm, subtitle[:120])


def draw_footer(pdf: canvas.Canvas, footer_text: str) -> None:
    width, _height = A4
    pdf.setStrokeColorRGB(0.75, 0.77, 0.81)
    pdf.line(18 * mm, 16 * mm, width - 18 * mm, 16 * mm)
    pdf.setFont("Helvetica", 7.5)
    pdf.setFillColorRGB(0.38, 0.42, 0.48)
    pdf.drawString(18 * mm, 11 * mm, footer_text[:140])


def draw_cover_page(pdf: canvas.Canvas, inputs: dict[str, Any], results: dict[str, Any], total_pages: int) -> None:
    width, height = A4
    draw_header(pdf, "Detailed Design Report - Vented Submersible Causeway", inputs["projectName"], 1, total_pages)
    pdf.setFillColorRGB(0.11, 0.14, 0.18)
    pdf.setFont("Helvetica-Bold", 22)
    pdf.drawCentredString(width / 2, height - 65 * mm, "STATUTORY DETAILED REPORT")
    pdf.setFont("Helvetica", 12)
    pdf.drawCentredString(width / 2, height - 75 * mm, inputs["projectName"])
    pdf.drawCentredString(width / 2, height - 83 * mm, f"Stream / Crossing: {inputs['streamName']}")
    pdf.drawCentredString(width / 2, height - 91 * mm, f"Location: {inputs['location']}")
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(24 * mm, height - 115 * mm, "Live imported values used for this report")
    pdf.setFont("Helvetica", 9.5)
    y = height - 125 * mm
    rows = [
        ("Date", inputs["date"]),
        ("Design discharge", f"{f(results['designDischarge'], 3)} m³/s"),
        ("Road top level", f"{f(inputs['rtl'], 3)} m"),
        ("Foundation level", f"{f(results['fbl'], 3)} m"),
        ("Ventway area", f"{f(results['aVent'], 3)} m²"),
        ("Afflux", f"{f(results['hAfflux'], 4)} m"),
        ("Maximum scour depth", f"{f(results['maxScourDepth'], 3)} m"),
        ("Live load type", inputs["liveLoadType"]),
    ]
    for label, value in rows:
        pdf.setFillColorRGB(0.25, 0.28, 0.33)
        pdf.drawString(28 * mm, y, label)
        pdf.setFillColorRGB(0.05, 0.07, 0.09)
        pdf.drawRightString(width - 28 * mm, y, str(value))
        pdf.setStrokeColorRGB(0.86, 0.88, 0.90)
        pdf.line(28 * mm, y - 2.5 * mm, width - 28 * mm, y - 2.5 * mm)
        y -= 8 * mm
    draw_footer(pdf, "Generated from spreadsheet variables using the statutory report generator")
    pdf.showPage()


def draw_summary_page(pdf: canvas.Canvas, inputs: dict[str, Any], results: dict[str, Any], total_pages: int, page_no: int) -> None:
    width, height = A4
    draw_header(pdf, "Hydraulic and Structural Summary", inputs["projectName"], page_no, total_pages)
    blocks = [
        ("Discharge", [
            ("Catchment area", f"{f(inputs['catchmentArea'], 3)} km²"),
            ("Rainfall intensity", f"{f(inputs['rainfallIntensity'], 3)} mm/hr"),
            ("Runoff coefficient", f(inputs["runoffCoefficient"], 3)),
            ("Governing method", results["governingMethod"]),
            ("Computed Q", f"{f(results['designDischarge'], 3)} m³/s"),
        ]),
        ("Hydraulic", [
            ("Vent area", f"{f(results['aVent'], 3)} m²"),
            ("Velocity at HFL", f"{f(results['velocityHFL'], 3)} m/s"),
            ("Obstruction at RTL", f"{f(results['pctObsRTL'], 2)} %"),
            ("Obstruction at HFL", f"{f(results['pctObsHFL'], 2)} %"),
            ("Afflux", f"{f(results['hAfflux'], 4)} m"),
        ]),
        ("Foundation", [
            ("Lacey perimeter", f"{f(results['laceyPerimeter'], 3)} m"),
            ("Scour depth", f"{f(results['maxScourDepth'], 3)} m"),
            ("Foundation level", f"{f(results['fbl'], 3)} m"),
            ("Recommended depth", f"{f(results['recommendedDepth'], 3)} m"),
            ("Scour safe", "YES" if results["scourSafe"] else "NO"),
        ]),
        ("Structural", [
            ("Self weight", f"{f(results['wSelf'], 3)} kN"),
            ("Silt load", f"{f(results['wSilt'], 3)} kN"),
            ("Live load", f"{f(results['wLive'], 3)} kN"),
            ("Uplift", f"{f(results['fUplift'], 3)} kN"),
            ("Total drag", f"{f(results['fDragTotal'], 3)} kN"),
        ]),
    ]
    x_positions = [18 * mm, 108 * mm]
    y_positions = [height - 48 * mm, height - 145 * mm]
    idx = 0
    for row in range(2):
        for col in range(2):
            title, items = blocks[idx]
            idx += 1
            x = x_positions[col]
            y_top = y_positions[row]
            pdf.setStrokeColorRGB(0.75, 0.79, 0.83)
            pdf.rect(x, y_top - 70 * mm, 75 * mm, 62 * mm, fill=0, stroke=1)
            pdf.setFillColorRGB(0.93, 0.95, 0.98)
            pdf.rect(x, y_top - 10 * mm, 75 * mm, 10 * mm, fill=1, stroke=0)
            pdf.setFillColorRGB(0.11, 0.14, 0.18)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(x + 4 * mm, y_top - 6.5 * mm, title)
            pdf.setFont("Helvetica", 8.5)
            y = y_top - 16 * mm
            for label, value in items:
                pdf.setFillColorRGB(0.25, 0.28, 0.33)
                pdf.drawString(x + 4 * mm, y, label)
                pdf.setFillColorRGB(0.05, 0.07, 0.09)
                pdf.drawRightString(x + 71 * mm, y, value)
                y -= 8.5 * mm
    draw_footer(pdf, "Summary page generated from current spreadsheet values")
    pdf.showPage()


def draw_command_page(pdf: canvas.Canvas, inputs: dict[str, Any], total_pages: int, page_no: int) -> None:
    width, height = A4
    draw_header(pdf, "Generator Notes and Usage", inputs["projectName"], page_no, total_pages)
    pdf.setFillColorRGB(0.08, 0.10, 0.13)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(20 * mm, height - 48 * mm, "How this detailed report is built")
    pdf.setFont("Helvetica", 9.5)
    paragraphs = [
        "This generator consumes the same spreadsheet keys used by the web app import page. CSV, XLSX, and JSON inputs are all accepted.",
        "The first pages are rebuilt from the live imported values and computed hydraulic or structural results. The remaining pages preserve the statutory source structure while carrying the current project identity in the page header and selected value substitutions in the body text.",
        "Use this output when you need a file-based A4 portrait statutory report rather than the shorter browser export.",
    ]
    y = height - 58 * mm
    for paragraph in paragraphs:
        wrapped = textwrap.wrap(paragraph, width=92)
        for line in wrapped:
            pdf.drawString(20 * mm, y, line)
            y -= 5.5 * mm
        y -= 4 * mm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(20 * mm, y - 3 * mm, "Command example")
    pdf.setFont("Courier", 8.5)
    command = "pnpm --filter @workspace/scripts run pdf:statutory -- --input /path/to/vars.xlsx --output /workspace/report.pdf"
    for line in textwrap.wrap(command, width=78):
        y -= 7 * mm
        pdf.drawString(24 * mm, y, line)
    draw_footer(pdf, "The full report continues with the detailed statutory pages")
    pdf.showPage()


def draw_text_page(
    pdf: canvas.Canvas,
    page_text: str,
    inputs: dict[str, Any],
    results: dict[str, Any],
    page_no: int,
    total_pages: int,
) -> None:
    title = "Detailed Statutory Pages"
    subtitle = f"{inputs['projectName']} | Stream: {inputs['streamName']} | Location: {inputs['location']}"
    draw_header(pdf, title, subtitle, page_no, total_pages)
    replacements = replacement_map(inputs, results)
    cleaned = " ".join((page_text or "").split())
    cleaned = apply_replacements(cleaned, replacements)
    if not cleaned:
        cleaned = "No extractable text was available on this source page. Page structure preserved for statutory page count continuity."
    pdf.setFont("Helvetica", 7.4)
    pdf.setFillColorRGB(0.08, 0.10, 0.13)
    wrapped_lines = textwrap.wrap(cleaned, width=115)
    y = A4[1] - 42 * mm
    bottom_limit = 20 * mm
    for line in wrapped_lines:
        if y < bottom_limit:
            break
        pdf.drawString(18 * mm, y, line)
        y -= 4.1 * mm
    draw_footer(pdf, f"Current design discharge {f(results['designDischarge'], 3)} m³/s | Foundation level {f(results['fbl'], 3)} m")
    pdf.showPage()


def build_pdf(input_path: Path | None, output_path: Path, source_pdf_path: Path) -> None:
    inputs = load_inputs(input_path)
    results = compute_results(inputs)
    reader = PdfReader(str(source_pdf_path))
    total_pages = len(reader.pages)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(output_path), pagesize=A4)
    pdf.setTitle("Submersible Bridge Design - Detailed Statutory Report")
    pdf.setAuthor("TRAE statutory generator")

    for page_index in range(total_pages):
        if page_index == 0:
            draw_cover_page(pdf, inputs, results, total_pages)
            continue
        if page_index == 1:
            draw_summary_page(pdf, inputs, results, total_pages, page_index + 1)
            continue
        if page_index == 2:
            draw_command_page(pdf, inputs, total_pages, page_index + 1)
            continue
        page_text = reader.pages[page_index].extract_text() or ""
        draw_text_page(pdf, page_text, inputs, results, page_index + 1, total_pages)

    pdf.save()


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a 169-page statutory A4 report from CSV/XLSX/JSON variables.")
    parser.add_argument("--input", dest="input_path", default=None, help="Path to CSV, XLSX, or JSON input variables file")
    parser.add_argument("--output", dest="output_path", default="/workspace/Submersible_Bridge_Design_Statutory_169p.pdf", help="Output PDF path")
    parser.add_argument(
        "--source-pdf",
        dest="source_pdf_path",
        default=str(Path(__file__).resolve().parents[2] / "attached_assets" / "Type Design of submersible causeway.pdf"),
        help="Source statutory PDF used as the page-count and narrative base",
    )
    argv = sys.argv[1:]
    if argv and argv[0] == "--":
        argv = argv[1:]
    args = parser.parse_args(argv)

    input_path = Path(args.input_path).resolve() if args.input_path else None
    output_path = Path(args.output_path).resolve()
    source_pdf_path = Path(args.source_pdf_path).resolve()
    build_pdf(input_path, output_path, source_pdf_path)
    print(f"Wrote statutory report to {output_path}")


if __name__ == "__main__":
    main()
